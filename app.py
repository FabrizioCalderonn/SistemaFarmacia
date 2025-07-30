from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import os
from datetime import datetime, timezone, timedelta
import tempfile
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import werkzeug
import csv

# Importar psycopg2 para PostgreSQL o sqlite3 como fallback
try:
    import psycopg2
    import psycopg2.extras
    DATABASE_TYPE = 'postgresql'
except ImportError:
    import sqlite3
    DATABASE_TYPE = 'sqlite'

# Configuración de farmacias
FARMACIAS = {
    'farmacia1': {
        'nombre': 'Farmacia Popular',
        'direccion': 'Calle 15 de Septiembre Bo. San Pedro No. 3, Metapán, Santa Ana',
        'telefono': '2442-0202'
    },
    'farmacia2': {
        'nombre': 'Farmacia El Angel',
        'direccion': '2a. Calle Oriente, Av. Ignacio Gomez, Bo. San Pedro #3, Metapán, Santa Ana',
        'telefono': '2442-0031'
    }
}

# Farmacia por defecto
FARMACIA_DEFAULT = 'farmacia1'

def get_local_datetime():
    """Obtener la fecha y hora local en formato ISO"""
    # Obtener la hora local (UTC-6 para El Salvador)
    local_tz = timezone(timedelta(hours=-6))
    return datetime.now(local_tz).isoformat()

app = Flask(__name__)
app.secret_key = 'tu_clave_secreta_aqui_12345'  # Necesario para las sesiones

# Configuración de la base de datos
if DATABASE_TYPE == 'postgresql':
    # Configuración para PostgreSQL (Supabase)
    DATABASE_URL = os.environ.get('DATABASE_URL')
    if DATABASE_URL and DATABASE_URL.startswith('postgres://'):
        DATABASE_URL = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
else:
    # Configuración para SQLite (local)
    DATABASE = 'farmacia.db'

def get_db_connection():
    """Obtener conexión a la base de datos"""
    if DATABASE_TYPE == 'postgresql':
        if not DATABASE_URL:
            raise Exception("DATABASE_URL no configurada")
        try:
            # Configurar timeout y reintentos
            return psycopg2.connect(
                DATABASE_URL,
                connect_timeout=10,
                options='-c statement_timeout=30000'
            )
        except Exception as e:
            print(f"Error de conexión a PostgreSQL: {e}")
            raise
    else:
        return sqlite3.connect(DATABASE)

def init_db():
    """Inicializar la base de datos y crear solo tabla de registros"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    if DATABASE_TYPE == 'postgresql':
        # Crear tabla de registros para PostgreSQL (inventario se lee desde CSV)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS registros (
                id SERIAL PRIMARY KEY,
                medicamento TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                precio REAL DEFAULT 0.0,
                fecha TEXT NOT NULL,
                fecha_ingreso TEXT NOT NULL,
                observaciones TEXT,
                laboratorio TEXT NOT NULL,
                tipo_movimiento TEXT DEFAULT 'VENTA',
                fecha_vencimiento TEXT,
                lote TEXT,
                medico TEXT,
                junta_vigilancia TEXT,
                numero_inscripcion_clinica TEXT,
                numero_factura TEXT,
                farmacia TEXT DEFAULT 'farmacia1'
            )
        ''')
        
        # Agregar columna numero_factura si no existe (para tablas existentes)
        try:
            cursor.execute('ALTER TABLE registros ADD COLUMN numero_factura TEXT')
        except:
            pass  # La columna ya existe
    else:
        # Crear tabla de registros para SQLite (inventario se lee desde CSV)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS registros (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                medicamento TEXT NOT NULL,
                cantidad INTEGER NOT NULL,
                precio REAL DEFAULT 0.0,
                fecha TEXT NOT NULL,
                fecha_ingreso TEXT NOT NULL,
                observaciones TEXT,
                laboratorio TEXT NOT NULL,
                tipo_movimiento TEXT DEFAULT 'VENTA',
                fecha_vencimiento TEXT,
                lote TEXT,
                medico TEXT,
                junta_vigilancia TEXT,
                numero_inscripcion_clinica TEXT,
                numero_factura TEXT,
                farmacia TEXT DEFAULT 'farmacia1'
            )
        ''')
        
        # Agregar columna numero_factura si no existe (para tablas existentes)
        try:
            cursor.execute('ALTER TABLE registros ADD COLUMN numero_factura TEXT')
        except:
            pass  # La columna ya existe
    
    conn.commit()
    conn.close()

def load_inventory_from_csv():
    """Cargar inventario desde archivo CSV usando el script dedicado"""
    if not os.path.exists('INVENTARIO PARA TRABAJO.csv'):
        print("Archivo 'INVENTARIO PARA TRABAJO.csv' no encontrado")
        return
    
    print("Función load_inventory_from_csv deshabilitada - archivo procesar_inventario.py eliminado")

def sincronizar_inventario_web():
    """Sincronizar inventario desde la aplicación web"""
    try:
        # Importar y ejecutar el script de sincronización simplificado
        from sincronizar_inventario_simple import sincronizar_inventario
        success, result = sincronizar_inventario()
        return success, result
        
    except Exception as e:
        print(f"Error al sincronizar inventario: {e}")
        return False, f"Error: {str(e)}"

def procesar_excel_web(archivo):
    """Procesar archivo Excel desde la aplicación web"""
    try:
        # Importar y ejecutar el script de procesamiento de Excel simplificado
        from procesar_excel_simple import procesar_excel_original, backup_antes_de_procesar
        
        # Crear backup antes de procesar
        if not backup_antes_de_procesar():
            return False, "Error al crear backup"
        
        # Guardar archivo temporalmente
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_archivo = f"excel_upload_{timestamp}.xlsx"
        ruta_archivo = os.path.join(tempfile.gettempdir(), nombre_archivo)
        archivo.save(ruta_archivo)
        
        # Procesar el archivo
        success, result = procesar_excel_original(ruta_archivo)
        
        # Limpiar archivo temporal
        try:
            os.remove(ruta_archivo)
        except:
            pass
        
        return success, result
        
    except Exception as e:
        print(f"Error al procesar Excel: {e}")
        return False, f"Error: {str(e)}"

def parse_inventory_file():
    """Parsear el archivo de inventario y extraer productos"""
    try:
        csv_path = 'INVENTARIO PARA TRABAJO.csv'
        if not os.path.exists(csv_path):
            return []
        
        productos = []
        with open(csv_path, 'r', encoding='utf-8-sig') as file:
            lines = file.readlines()
            
            # Buscar la línea de encabezados
            header_found = False
            for i, line in enumerate(lines):
                line = line.strip()
                
                # Detectar línea de encabezados (más flexible)
                if 'Codigo' in line and 'Nombre' in line:
                    header_found = True
                    print(f"Encabezados encontrados en línea {i+1}: {line}")
                    continue
                
                # Procesar líneas de datos después de encontrar encabezados
                if header_found and line and not line.startswith('"Listado') and not line.startswith('"Reportes'):
                    # Limpiar la línea de comillas extra y espacios
                    line = line.replace('"', '')
                    # Dividir por tabulaciones y limpiar espacios extra
                    parts = [part.strip() for part in line.split('\t') if part.strip()]
                    
                    if len(parts) >= 3:  # Cambiado de 4 a 3
                        codigo = parts[0]
                        nombre = parts[1]
                        modelo_marca = parts[2]  # Modelo y marca están juntos
                        
                        # Extraer marca/laboratorio del tercer campo
                        # Buscar la última coma para separar modelo de marca
                        if ',' in modelo_marca:
                            modelo, marca = modelo_marca.rsplit(',', 1)
                            modelo = modelo.strip()
                            marca = marca.strip()
                        else:
                            modelo = modelo_marca.strip()
                            marca = 'Sin especificar'
                        
                        # Solo agregar si tiene código y nombre válidos (no solo espacios)
                        if codigo and nombre and len(codigo.strip()) > 3 and len(nombre.strip()) > 3:
                            productos.append({
                                'codigo': codigo.strip(),
                                'nombre': nombre.strip(),
                                'modelo': modelo,
                                'marca': marca,
                                'laboratorio': marca if marca else 'Sin especificar'
                            })
                        else:
                            print(f"Línea descartada - código: '{codigo}', nombre: '{nombre}'")
                    else:
                        print(f"Línea con menos de 3 partes: {parts}")
        
        print(f"Productos parseados: {len(productos)}")
        return productos
    except Exception as e:
        print(f"Error al parsear inventario: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_laboratorios():
    """Obtener lista de laboratorios únicos desde archivo de inventario"""
    try:
        # Usar el módulo simple
        from inventario_simple import get_laboratorios as get_labs
        return get_labs()
    except Exception as e:
        print(f"Error al obtener laboratorios: {e}")
        import traceback
        traceback.print_exc()
        return []

def get_medicamentos_by_laboratorio(laboratorio):
    """Obtener medicamentos por laboratorio desde archivo de inventario"""
    try:
        # Usar el módulo simple
        from inventario_simple import get_medicamentos_by_laboratorio as get_meds_simple
        print(f"App.py: Llamando a get_medicamentos_by_laboratorio con '{laboratorio}'")
        result = get_meds_simple(laboratorio)
        print(f"App.py: Resultado obtenido: {len(result)} medicamentos")
        return result
    except Exception as e:
        print(f"Error al obtener medicamentos: {e}")
        return []

def buscar_productos(termino):
    """Buscar productos por nombre o presentación desde CSV"""
    try:
        if not os.path.exists('INVENTARIO PARA TRABAJO.csv'):
            return []
        
        termino = termino.lower()
        productos = []
        with open('INVENTARIO PARA TRABAJO.csv', 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                medicamento = row.get('medicamento', '').lower()
                presentacion = row.get('presentacion', '').lower()
                
                if termino in medicamento or termino in presentacion:
                    productos.append({
                        'laboratorio': row.get('laboratorio', ''),
                        'medicamento': row.get('medicamento', ''),
                        'presentacion': row.get('presentacion', ''),
                        'precio': float(row.get('precio', 0)),
                        'stock': int(row.get('stock', 0))
                    })
        
        return productos
    except Exception as e:
        print(f"Error al buscar productos: {e}")
        return []

def get_estadisticas():
    """Obtener estadísticas del inventario desde CSV"""
    try:
        if not os.path.exists('INVENTARIO PARA TRABAJO.csv'):
            return {
                'total_productos': 0,
                'total_laboratorios': 0,
                'stock_bajo': 0,
                'sin_stock': 0
            }
        
        total_productos = 0
        laboratorios = set()
        stock_bajo = 0
        sin_stock = 0
        
        with open('INVENTARIO PARA TRABAJO.csv', 'r', encoding='utf-8') as file:
            reader = csv.DictReader(file)
            for row in reader:
                total_productos += 1
                if row.get('laboratorio'):
                    laboratorios.add(row['laboratorio'])
                
                stock = int(row.get('stock', 0))
                if stock == 0:
                    sin_stock += 1
                elif stock < 10:
                    stock_bajo += 1
        
        return {
            'total_productos': total_productos,
            'total_laboratorios': len(laboratorios),
            'stock_bajo': stock_bajo,
            'sin_stock': sin_stock
        }
    except Exception as e:
        print(f"Error al obtener estadísticas: {e}")
        return {
            'total_productos': 0,
            'total_laboratorios': 0,
            'stock_bajo': 0,
            'sin_stock': 0
        }

def get_farmacia_actual():
    """Obtener la farmacia actual desde la sesión o por defecto"""
    return session.get('farmacia_actual', FARMACIA_DEFAULT)

def get_farmacia_info(farmacia_id):
    """Obtener información de una farmacia específica"""
    return FARMACIAS.get(farmacia_id, FARMACIAS[FARMACIA_DEFAULT])

# Inicializar base de datos (solo para registros, no inventario)
init_db()

@app.route('/')
def index():
    """Página principal"""
    # Inicializar base de datos si es necesario
    initialize_database()
    
    # Verificar que el inventario esté cargado
    try:
        from inventario_simple import cargar_inventario
        cargar_inventario()
        print("Inventario cargado en la página principal")
    except Exception as e:
        print(f"Error al cargar inventario en página principal: {e}")
    
    laboratorios = get_laboratorios()
    estadisticas = get_estadisticas()
    farmacia_actual = get_farmacia_actual()
    farmacia_info = get_farmacia_info(farmacia_actual)
    
    return render_template('index.html', 
                         laboratorios=laboratorios, 
                         estadisticas=estadisticas,
                         farmacia_actual=farmacia_actual,
                         farmacia_info=farmacia_info)

@app.route('/api/laboratorios')
def api_laboratorios():
    """API para obtener lista de laboratorios"""
    # Verificar que el inventario esté cargado
    try:
        from inventario_simple import cargar_inventario
        cargar_inventario()
        print("Inventario cargado en API laboratorios")
    except Exception as e:
        print(f"Error al cargar inventario en API laboratorios: {e}")
    
    laboratorios = get_laboratorios()
    print(f"API: Laboratorios disponibles: {laboratorios}")
    return jsonify(laboratorios)

@app.route('/api/medicamentos/<laboratorio>')
def api_medicamentos(laboratorio):
    """API para obtener medicamentos por laboratorio"""
    print(f"=== API MEDICAMENTOS ===")
    print(f"API: Buscando medicamentos para laboratorio: '{laboratorio}'")
    print(f"API: Tipo de laboratorio: {type(laboratorio)}")
    print(f"API: Longitud del laboratorio: {len(laboratorio)}")
    
    # Verificar que el inventario esté cargado
    try:
        from inventario_simple import cargar_inventario
        cargar_inventario()
        print("Inventario cargado en API medicamentos")
    except Exception as e:
        print(f"Error al cargar inventario en API medicamentos: {e}")
    
    # Decodificar la URL si es necesario
    import urllib.parse
    laboratorio_decoded = urllib.parse.unquote(laboratorio)
    print(f"API: Laboratorio decodificado: '{laboratorio_decoded}'")
    
    try:
        medicamentos = get_medicamentos_by_laboratorio(laboratorio_decoded)
        print(f"API: Encontrados {len(medicamentos)} medicamentos")
        print(f"API: Tipo de resultado: {type(medicamentos)}")
        
        if medicamentos:
            print(f"API: Primer medicamento: {medicamentos[0]}")
        
        return jsonify(medicamentos)
    except Exception as e:
        print(f"API: Error en api_medicamentos: {e}")
        import traceback
        traceback.print_exc()
        return jsonify([])

@app.route('/api/buscar_simple')
def api_buscar_simple():
    """API de búsqueda simplificada para debug"""
    termino = request.args.get('q', '')
    print(f"Búsqueda simple solicitada: '{termino}'")
    
    if len(termino) < 2:
        return jsonify([])
    
    # Usar directamente la función de búsqueda
    from inventario_simple import buscar_productos
    productos = buscar_productos(termino)
    
    # Convertir de forma más simple
    resultado = []
    for producto in productos:
        resultado.append([
            producto.get('laboratorio', ''),
            producto.get('medicamento', ''),
            producto.get('presentacion', '')
        ])
    
    print(f"Resultado simple: {resultado}")
    return jsonify(resultado)

@app.route('/api/buscar')
def api_buscar():
    """API para buscar productos"""
    termino = request.args.get('q', '')
    print(f"Búsqueda solicitada: '{termino}'")
    
    if len(termino) < 2:
        print("Término muy corto, devolviendo lista vacía")
        return jsonify([])
    
    productos = buscar_productos(termino)
    print(f"Productos encontrados (formato original): {productos}")
    
    # Convertir al formato esperado por el frontend: [laboratorio, medicamento, presentacion]
    productos_formateados = []
    print(f"Procesando {len(productos)} productos...")
    
    for i, producto in enumerate(productos):
        print(f"Producto {i+1}: {producto}")
        try:
            productos_formateados.append([
                producto['laboratorio'],
                producto['medicamento'],
                producto['presentacion']
            ])
            print(f"Producto {i+1} convertido exitosamente")
        except Exception as e:
            print(f"Error convirtiendo producto {i+1}: {e}")
            print(f"Tipo de producto: {type(producto)}")
            print(f"Claves disponibles: {producto.keys() if hasattr(producto, 'keys') else 'No es un diccionario'}")
    
    print(f"Productos formateados: {productos_formateados}")
    print(f"Total productos formateados: {len(productos_formateados)}")
    
    # Verificar que no esté vacío
    if not productos_formateados:
        print("ADVERTENCIA: productos_formateados está vacío!")
    
    return jsonify(productos_formateados)

@app.route('/guardar_registro', methods=['POST'])
def guardar_registro():
    """Guardar un nuevo registro"""
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['laboratorio', 'medicamento', 'cantidad', 'fecha']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Campo {field} es requerido'})
        
        # Insertar registro
        conn = get_db_connection()
        cursor = conn.cursor()
        
        farmacia_actual = get_farmacia_actual()
        
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('''
                INSERT INTO registros (laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones, tipo_movimiento, 
                                     fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica, numero_factura, farmacia)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                data['laboratorio'],
                data['medicamento'],
                data['cantidad'],
                data['fecha'],
                get_local_datetime(),
                data.get('observaciones', ''),
                data.get('tipo_movimiento', 'VENTA'),
                data.get('fecha_vencimiento', ''),
                data.get('lote', ''),
                data.get('medico', ''),
                data.get('junta_vigilancia', ''),
                data.get('numero_inscripcion_clinica', ''),
                data.get('numero_factura', ''),
                farmacia_actual
            ))
            

        else:
            cursor.execute('''
                INSERT INTO registros (laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones, tipo_movimiento, 
                                 fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica, numero_factura, farmacia)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                data['laboratorio'],
                data['medicamento'],
                data['cantidad'],
                data['fecha'],
                get_local_datetime(),
                data.get('observaciones', ''),
                data.get('tipo_movimiento', 'VENTA'),
                data.get('fecha_vencimiento', ''),
                data.get('lote', ''),
                data.get('medico', ''),
                data.get('junta_vigilancia', ''),
                data.get('numero_inscripcion_clinica', ''),
                data.get('numero_factura', ''),
                farmacia_actual
            ))
            

        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Registrado exitosamente'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/historial')
def historial():
    """Página de historial de registros"""
    return render_template('historial.html')

@app.route('/api/registros')
def api_registros():
    """API para obtener registros con filtros"""
    try:
        # Parámetros de filtro
        fecha_inicio = request.args.get('fecha_inicio', '')
        fecha_fin = request.args.get('fecha_fin', '')
        laboratorio = request.args.get('laboratorio', '')
        medicamento = request.args.get('medicamento', '')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener farmacia actual
        farmacia_actual = get_farmacia_actual()
        
        # Construir consulta con filtros
        query = '''
            SELECT id, laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones,
                   fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica, numero_factura
            FROM registros 
            WHERE farmacia = %s
        '''
        params = [farmacia_actual]
        
        if fecha_inicio:
            query += ' AND fecha_ingreso >= %s'
            params.append(fecha_inicio + 'T00:00:00')
        
        if fecha_fin:
            query += ' AND fecha_ingreso <= %s'
            params.append(fecha_fin + 'T23:59:59')
        
        if laboratorio:
            query += ' AND laboratorio LIKE %s'
            params.append(f'%{laboratorio}%')
        
        if medicamento:
            query += ' AND medicamento LIKE %s'
            params.append(f'%{medicamento}%')
        
        query += ' ORDER BY fecha_ingreso DESC LIMIT 1000'
        
        cursor.execute(query, params)
        registros = cursor.fetchall()
        conn.close()
        
        # Formatear registros
        registros_formateados = []
        for reg in registros:
            registros_formateados.append({
                'id': reg[0],
                'laboratorio': reg[1],
                'medicamento': reg[2],
                'cantidad': reg[3],
                'fecha': reg[4],
                'fecha_ingreso': reg[5],
                'observaciones': reg[6],
                'fecha_vencimiento': reg[7] if len(reg) > 7 else '',
                'lote': reg[8] if len(reg) > 8 else '',
                'medico': reg[9] if len(reg) > 9 else '',
                'junta_vigilancia': reg[10] if len(reg) > 10 else '',
                'numero_inscripcion_clinica': reg[11] if len(reg) > 11 else '',
                'numero_factura': reg[12] if len(reg) > 12 else ''
            })
        
        return jsonify(registros_formateados)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/api/registros/<int:registro_id>', methods=['PUT'])
def actualizar_registro(registro_id):
    """Actualizar un registro existente"""
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['laboratorio', 'medicamento', 'cantidad', 'fecha']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'success': False, 'message': f'Campo {field} es requerido'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el registro existe
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('SELECT id FROM registros WHERE id = %s', (registro_id,))
        else:
            cursor.execute('SELECT id FROM registros WHERE id = ?', (registro_id,))
        registro_existe = cursor.fetchone()
        
        if not registro_existe:
            conn.close()
            return jsonify({'success': False, 'message': 'Registro no encontrado'})
        
        # Actualizar el registro (sin modificar inventario - se maneja desde CSV)
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('''
                UPDATE registros 
                SET laboratorio = %s, medicamento = %s, cantidad = %s, fecha = %s, 
                    observaciones = %s, fecha_vencimiento = %s, lote = %s, medico = %s, junta_vigilancia = %s, numero_inscripcion_clinica = %s, numero_factura = %s
                WHERE id = %s
            ''', (
                data['laboratorio'],
                data['medicamento'],
                data['cantidad'],
                data['fecha'],
                data.get('observaciones', ''),
                data.get('fecha_vencimiento', ''),
                data.get('lote', ''),
                data.get('medico', ''),
                data.get('junta_vigilancia', ''),
                data.get('numero_inscripcion_clinica', ''),
                data.get('numero_factura', ''),
                registro_id
            ))
        else:
            cursor.execute('''
                UPDATE registros 
                SET laboratorio = ?, medicamento = ?, cantidad = ?, fecha = ?, 
                    observaciones = ?, fecha_vencimiento = ?, lote = ?, medico = ?, junta_vigilancia = ?, numero_inscripcion_clinica = ?, numero_factura = ?
                WHERE id = ?
            ''', (
                data['laboratorio'],
                data['medicamento'],
                data['cantidad'],
                data['fecha'],
                data.get('observaciones', ''),
                data.get('fecha_vencimiento', ''),
                data.get('lote', ''),
                data.get('medico', ''),
                data.get('junta_vigilancia', ''),
                data.get('numero_inscripcion_clinica', ''),
                data.get('numero_factura', ''),
                registro_id
            ))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Registro actualizado exitosamente'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/api/registros/<int:registro_id>', methods=['DELETE'])
def eliminar_registro(registro_id):
    """Eliminar un registro"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar que el registro existe
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('SELECT id FROM registros WHERE id = %s', (registro_id,))
        else:
            cursor.execute('SELECT id FROM registros WHERE id = ?', (registro_id,))
        registro = cursor.fetchone()
        
        if not registro:
            conn.close()
            return jsonify({'success': False, 'message': 'Registro no encontrado'})
        
        # Eliminar el registro (sin modificar inventario - se maneja desde CSV)
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('DELETE FROM registros WHERE id = %s', (registro_id,))
        else:
            cursor.execute('DELETE FROM registros WHERE id = ?', (registro_id,))
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'success': True, 
            'message': 'Registro eliminado exitosamente'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/exportar_excel')
def exportar_excel():
    """Exportar registros a Excel"""
    try:
        # Obtener farmacia actual
        farmacia_actual = get_farmacia_actual()
        
        # Obtener registros
        conn = get_db_connection()
        cursor = conn.cursor()
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('''
                SELECT laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones,
                       fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica
                FROM registros 
                WHERE farmacia = %s
                ORDER BY fecha_ingreso DESC
            ''', (farmacia_actual,))
        else:
            cursor.execute('''
                SELECT laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones,
                       fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica
                FROM registros 
                WHERE farmacia = ?
                ORDER BY fecha_ingreso DESC
            ''', (farmacia_actual,))
        registros = cursor.fetchall()
        conn.close()
        
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Farmacia"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Encabezados
        headers = ['Fecha Ingreso', 'Medicamento', 'Laboratorio', 'Cantidad', 'Fecha Compra', 'Fecha Vencimiento', 'Lote', 'Médico', 'Junta Vigilancia', 'Número Inscripción Clínica', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        for row, registro in enumerate(registros, 2):
            # Formatear fecha de ingreso para que sea más legible
            fecha_ingreso = registro[4]
            if fecha_ingreso:
                try:
                    fecha_obj = datetime.fromisoformat(fecha_ingreso.replace('Z', '+00:00'))
                    fecha_formateada = fecha_obj.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    fecha_formateada = fecha_ingreso
            else:
                fecha_formateada = ''
            
            ws.cell(row=row, column=1, value=fecha_formateada)  # Fecha Ingreso
            ws.cell(row=row, column=2, value=registro[1])  # Medicamento
            ws.cell(row=row, column=3, value=registro[0])  # Laboratorio
            ws.cell(row=row, column=4, value=registro[2])  # Cantidad
            ws.cell(row=row, column=5, value=registro[3])  # Fecha Compra
            ws.cell(row=row, column=6, value=registro[6] if len(registro) > 6 else '')  # Fecha Vencimiento
            ws.cell(row=row, column=7, value=registro[7] if len(registro) > 7 else '')  # Lote
            ws.cell(row=row, column=8, value=registro[8] if len(registro) > 8 else '')  # Médico
            ws.cell(row=row, column=9, value=registro[9] if len(registro) > 9 else '')  # Junta Vigilancia
            ws.cell(row=row, column=10, value=registro[10] if len(registro) > 10 else '')  # Número Inscripción Clínica
            ws.cell(row=row, column=11, value=registro[5])  # Observaciones

        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"registros_farmacia_{timestamp}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/inventario')
def inventario():
    """Página de inventario"""
    return render_template('inventario.html')

@app.route('/api/inventario')
def api_inventario():
    """API para obtener inventario con filtros"""
    try:
        laboratorio = request.args.get('laboratorio', '')
        stock_bajo = request.args.get('stock_bajo', 'false') == 'true'
        sin_stock = request.args.get('sin_stock', 'false') == 'true'
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        query = 'SELECT laboratorio, medicamento, presentacion, precio, stock FROM inventario WHERE 1=1'
        params = []
        
        if laboratorio:
            query += ' AND laboratorio LIKE %s'
            params.append(f'%{laboratorio}%')
        
        if stock_bajo:
            query += ' AND stock < 10 AND stock > 0'
        
        if sin_stock:
            query += ' AND stock = 0'
        
        query += ' ORDER BY laboratorio, medicamento LIMIT 1000'
        
        cursor.execute(query, params)
        inventario = cursor.fetchall()
        conn.close()
        
        inventario_formateado = []
        for item in inventario:
            inventario_formateado.append({
                'laboratorio': item[0],
                'medicamento': item[1],
                'presentacion': item[2],
                'precio': item[3],
                'stock': item[4]
            })
        
        return jsonify(inventario_formateado)
        
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/sincronizar_inventario', methods=['POST'])
def sincronizar_inventario_route():
    """Ruta para sincronizar inventario desde la web"""
    try:
        # Crear backup antes de sincronizar
        from sincronizar_inventario_simple import backup_registros
        if not backup_registros():
            return jsonify({'success': False, 'message': 'Error al crear backup'})
        
        # Realizar sincronización
        success, result = sincronizar_inventario_web()
        if success:
            return jsonify({
                'success': True, 
                'message': 'Inventario sincronizado exitosamente',
                'estadisticas': result
            })
        else:
            return jsonify({'success': False, 'message': result})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/eliminar_productos_obsoletos', methods=['POST'])
def eliminar_productos_obsoletos_route():
    """Ruta para eliminar productos obsoletos desde la web"""
    try:
        # Crear backup antes de eliminar
        from eliminar_productos_simple import backup_antes_de_eliminar
        if not backup_antes_de_eliminar():
            return jsonify({'success': False, 'message': 'Error al crear backup'})
        
        # Realizar eliminación
        from eliminar_productos_simple import eliminar_productos_obsoletos
        success, result = eliminar_productos_obsoletos()
        
        if success:
            if isinstance(result, dict):
                return jsonify({
                    'success': True, 
                    'message': 'Productos obsoletos eliminados exitosamente',
                    'estadisticas': result
                })
            else:
                return jsonify({
                    'success': True, 
                    'message': result
                })
        else:
            return jsonify({'success': False, 'message': result})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/cargar_excel', methods=['POST'])
def cargar_excel_route():
    """Ruta para cargar archivo Excel desde la web"""
    try:
        # Verificar si se envió un archivo
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        archivo = request.files['archivo']
        
        # Verificar si el archivo está vacío
        if archivo.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        # Verificar extensión del archivo
        if not archivo.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'success': False, 'message': 'Solo se permiten archivos Excel (.xlsx, .xls)'})
        
        # Procesar el archivo Excel
        success, result = procesar_excel_web(archivo)
        
        if success:
            return jsonify({
                'success': True, 
                'message': 'Archivo Excel procesado exitosamente',
                'estadisticas': result
            })
        else:
            return jsonify({'success': False, 'message': result})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/actualizar_inventario_csv', methods=['POST'])
def actualizar_inventario_csv():
    """Actualizar el archivo CSV de inventario desde la web"""
    try:
        # Verificar si se envió un archivo
        if 'archivo' not in request.files:
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        archivo = request.files['archivo']
        
        # Verificar si el archivo está vacío
        if archivo.filename == '':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        
        # Verificar extensión del archivo
        if not archivo.filename.lower().endswith('.csv'):
            return jsonify({'success': False, 'message': 'Solo se permiten archivos CSV'})
        
        # Crear backup del CSV actual
        if os.path.exists('INVENTARIO PARA TRABAJO.csv'):
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = f"INVENTARIO PARA TRABAJO_backup_{timestamp}.csv"
            os.rename('INVENTARIO PARA TRABAJO.csv', backup_file)
        
        # Guardar el nuevo archivo
        archivo.save('INVENTARIO PARA TRABAJO.csv')
        
        return jsonify({
            'success': True, 
            'message': 'Inventario CSV actualizado exitosamente'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/cambiar_farmacia', methods=['GET', 'POST'])
def cambiar_farmacia():
    """Página para cambiar de farmacia"""
    if request.method == 'POST':
        farmacia = request.args.get('farmacia', FARMACIA_DEFAULT)
        if farmacia in FARMACIAS:
            # Guardar la farmacia en la sesión
            session['farmacia_actual'] = farmacia
            return jsonify({'success': True, 'message': f'Farmacia cambiada a {FARMACIAS[farmacia]["nombre"]}'})
        else:
            return jsonify({'success': False, 'message': 'Farmacia no válida'})
    
    farmacias = FARMACIAS
    farmacia_actual = get_farmacia_actual()
    return render_template('cambiar_farmacia.html', farmacias=farmacias, farmacia_actual=farmacia_actual)

@app.route('/api/farmacias')
def api_farmacias():
    """API para obtener lista de farmacias"""
    return jsonify(FARMACIAS)

@app.route('/api/farmacia_actual')
def api_farmacia_actual():
    """API para obtener la farmacia actual"""
    farmacia_actual = get_farmacia_actual()
    return jsonify({
        'farmacia': farmacia_actual,
        'info': get_farmacia_info(farmacia_actual)
    })

@app.route('/debug/buscar/<termino>')
def debug_buscar(termino):
    """Debug específico para la función de búsqueda"""
    try:
        from inventario_simple import buscar_productos
        
        result = {
            'termino_buscado': termino,
            'productos_encontrados': [],
            'total_encontrados': 0,
            'error': None
        }
        
        # Probar búsqueda
        try:
            productos = buscar_productos(termino)
            result['productos_encontrados'] = productos
            result['total_encontrados'] = len(productos)
        except Exception as e:
            result['error'] = str(e)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/debug/busqueda')
def debug_busqueda():
    """Debug específico para las funciones de búsqueda"""
    try:
        from inventario_simple import get_laboratorios, get_medicamentos_by_laboratorio, buscar_productos
        
        result = {
            'laboratorios_disponibles': [],
            'medicamentos_por_laboratorio': {},
            'busqueda_ejemplo': {},
            'error': None
        }
        
        # Obtener laboratorios
        try:
            laboratorios = get_laboratorios()
            result['laboratorios_disponibles'] = laboratorios
            
            # Probar obtener medicamentos para cada laboratorio
            for lab in laboratorios[:3]:  # Solo los primeros 3
                medicamentos = get_medicamentos_by_laboratorio(lab)
                result['medicamentos_por_laboratorio'][lab] = {
                    'total': len(medicamentos),
                    'primeros_5': medicamentos[:5]
                }
        except Exception as e:
            result['error_laboratorios'] = str(e)
        
        # Probar búsqueda
        try:
            busqueda_result = buscar_productos('MARSCHALL')
            result['busqueda_ejemplo'] = {
                'termino': 'MARSCHALL',
                'total_encontrados': len(busqueda_result),
                'primeros_3': busqueda_result[:3]
            }
        except Exception as e:
            result['error_busqueda'] = str(e)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/debug/inventario')
def debug_inventario():
    """Debug específico para el módulo de inventario simple"""
    try:
        from inventario_simple import cargar_inventario, get_laboratorios, get_estadisticas, debug_inventario as debug_inv
        
        result = {
            'modulo_cargado': True,
            'inventario_cargado': False,
            'laboratorios': [],
            'estadisticas': {},
            'debug_detallado': {},
            'error': None
        }
        
        # Intentar cargar inventario
        try:
            cargar_inventario()
            result['inventario_cargado'] = True
        except Exception as e:
            result['error_carga'] = str(e)
        
        # Obtener laboratorios
        try:
            result['laboratorios'] = get_laboratorios()
        except Exception as e:
            result['error_laboratorios'] = str(e)
        
        # Obtener estadísticas
        try:
            result['estadisticas'] = get_estadisticas()
        except Exception as e:
            result['error_estadisticas'] = str(e)
        
        # Debug detallado
        try:
            result['debug_detallado'] = debug_inv()
        except Exception as e:
            result['error_debug'] = str(e)
        
        # Análisis de formato
        try:
            from inventario_simple import analizar_formato_linea
            result['analisis_formato'] = analizar_formato_linea()
        except Exception as e:
            result['error_analisis'] = str(e)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({
            'modulo_cargado': False,
            'error': str(e)
        })

@app.route('/debug/analyze')
def debug_analyze():
    """Analizar el archivo línea por línea para debug"""
    try:
        csv_path = 'INVENTARIO PARA TRABAJO.csv'
        result = {
            'archivo_existe': os.path.exists(csv_path),
            'analisis_lineas': []
        }
        
        if os.path.exists(csv_path):
            with open(csv_path, 'r', encoding='utf-8-sig') as file:
                lines = file.readlines()
                
                for i, line in enumerate(lines[:20]):  # Solo las primeras 20 líneas
                    line_clean = line.strip()
                    if line_clean:
                        parts = [part.strip() for part in line_clean.split('\t') if part.strip()]
                        result['analisis_lineas'].append({
                            'numero': i + 1,
                            'linea_original': line_clean,
                            'partes': parts,
                            'num_partes': len(parts),
                            'tiene_codigo': 'Codigo' in line_clean,
                            'tiene_nombre': 'Nombre' in line_clean
                        })
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/debug/csv')
def debug_csv():
    """Ruta de debug para verificar el archivo CSV"""
    try:
        csv_path = 'INVENTARIO PARA TRABAJO.csv'
        result = {
            'archivo_existe': os.path.exists(csv_path),
            'ruta_absoluta': os.path.abspath(csv_path),
            'archivos_en_directorio': os.listdir('.'),
            'tamaño_archivo': 0,
            'laboratorios': [],
            'productos_parseados': 0,
            'error': None
        }
        
        if os.path.exists(csv_path):
            result['tamaño_archivo'] = os.path.getsize(csv_path)
            
            # Usar la nueva función de parsing
            productos = parse_inventory_file()
            result['productos_parseados'] = len(productos)
            
            # Extraer laboratorios únicos
            laboratorios = set()
            for producto in productos:
                if producto['laboratorio'] and producto['laboratorio'] != 'Sin especificar':
                    laboratorios.add(producto['laboratorio'])
            
            result['laboratorios'] = sorted(list(laboratorios))
            
            # Mostrar algunos productos de ejemplo
            if productos:
                result['ejemplo_productos'] = productos[:3]
            
            # También mostrar las primeras líneas del archivo
            try:
                with open(csv_path, 'r', encoding='utf-8-sig') as file:
                    first_lines = [file.readline() for _ in range(10)]
                    result['primeras_lineas'] = first_lines
            except Exception as e:
                result['error_lectura'] = str(e)
        
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)})

@app.route('/debug/med_pharma')
def debug_med_pharma():
    """Debug específico para med pharma"""
    try:
        from inventario_simple import debug_inventario, cargar_inventario
        
        # Cargar inventario
        cargar_inventario()
        
        # Obtener información de debug
        debug_info = debug_inventario()
        
        # Buscar específicamente productos de "med pharma"
        med_pharma_productos = []
        from inventario_simple import INVENTARIO_MEMORIA
        
        for producto in INVENTARIO_MEMORIA:
            if 'med pharma' in producto['laboratorio'].lower():
                med_pharma_productos.append({
                    'nombre': producto['nombre'],
                    'laboratorio': producto['laboratorio'],
                    'modelo': producto['modelo']
                })
        
        return jsonify({
            'debug_info': debug_info,
            'med_pharma_productos': med_pharma_productos,
            'total_med_pharma': len(med_pharma_productos)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)})

# Para Vercel
app.debug = False

# Inicializar base de datos solo cuando sea necesario
def initialize_database():
    """Inicializar la base de datos solo cuando sea necesario"""
    try:
        init_db()
        print("Base de datos inicializada correctamente")
        return True
    except Exception as e:
        print(f"Error al inicializar la base de datos: {e}")
        return False 