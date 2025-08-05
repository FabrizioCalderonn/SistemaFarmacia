#!/usr/bin/env python3
"""
Script para probar la exportación de Excel y verificar la estructura de los datos
"""

import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import tempfile
import os

def insert_test_data():
    """Insertar datos de prueba para verificar el Excel"""
    print("📝 Insertando datos de prueba...")
    
    try:
        conn = sqlite3.connect('farmacia.db')
        cursor = conn.cursor()
        
        # Datos de prueba para venta
        venta_data = {
            'laboratorio': 'Laboratorio Pfizer',
            'medicamento': 'Paracetamol 500mg',
            'cantidad': 10,
            'fecha': '2024-01-15',
            'fecha_ingreso': datetime.now().isoformat(),
            'observaciones': 'Venta normal al público',
            'tipo_movimiento': 'VENTA',
            'fecha_vencimiento': '2025-12-31',
            'lote': 'LOTE2024001',
            'medico': 'Dr. Juan Pérez',
            'junta_vigilancia': 'JV001234',
            'numero_inscripcion_clinica': 'NIC987654',
            'numero_factura': 'FAC-2024-001',
            'codigo_empleado': 'EMP001',
            'farmacia': 'farmacia1'
        }
        
        # Datos de prueba para devolución
        devolucion_data = {
            'laboratorio': 'Laboratorio Bayer',
            'medicamento': 'Aspirina 100mg',
            'cantidad': 5,
            'fecha': '2024-01-16',
            'fecha_ingreso': datetime.now().isoformat(),
            'observaciones': 'Devolución - producto defectuoso',
            'tipo_movimiento': 'DEVOLUCION',
            'fecha_vencimiento': '2025-06-30',
            'lote': 'LOTE2024002',
            'medico': '',  # No requerido para devoluciones
            'junta_vigilancia': '',  # No requerido para devoluciones
            'numero_inscripcion_clinica': '',  # No requerido para devoluciones
            'numero_factura': 'DEV-2024-001',
            'codigo_empleado': 'EMP002',
            'farmacia': 'farmacia1'
        }
        
        # Insertar venta
        cursor.execute('''
            INSERT INTO registros (laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones, tipo_movimiento, 
                                 fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica, numero_factura, codigo_empleado, farmacia)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            venta_data['laboratorio'],
            venta_data['medicamento'],
            venta_data['cantidad'],
            venta_data['fecha'],
            venta_data['fecha_ingreso'],
            venta_data['observaciones'],
            venta_data['tipo_movimiento'],
            venta_data['fecha_vencimiento'],
            venta_data['lote'],
            venta_data['medico'],
            venta_data['junta_vigilancia'],
            venta_data['numero_inscripcion_clinica'],
            venta_data['numero_factura'],
            venta_data['codigo_empleado'],
            venta_data['farmacia']
        ))
        
        # Insertar devolución
        cursor.execute('''
            INSERT INTO registros (laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones, tipo_movimiento, 
                                 fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica, numero_factura, codigo_empleado, farmacia)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            devolucion_data['laboratorio'],
            devolucion_data['medicamento'],
            devolucion_data['cantidad'],
            devolucion_data['fecha'],
            devolucion_data['fecha_ingreso'],
            devolucion_data['observaciones'],
            devolucion_data['tipo_movimiento'],
            devolucion_data['fecha_vencimiento'],
            devolucion_data['lote'],
            devolucion_data['medico'],
            devolucion_data['junta_vigilancia'],
            devolucion_data['numero_inscripcion_clinica'],
            devolucion_data['numero_factura'],
            devolucion_data['codigo_empleado'],
            devolucion_data['farmacia']
        ))
        
        conn.commit()
        conn.close()
        print("✅ Datos de prueba insertados correctamente")
        return True
        
    except Exception as e:
        print(f"❌ Error al insertar datos de prueba: {e}")
        return False

def test_excel_export():
    """Probar la exportación de Excel"""
    print("\n📊 Probando exportación de Excel...")
    
    try:
        conn = sqlite3.connect('farmacia.db')
        cursor = conn.cursor()
        
        # Obtener registros
        cursor.execute('''
            SELECT laboratorio, medicamento, cantidad, fecha, fecha_ingreso, observaciones,
                   fecha_vencimiento, lote, medico, junta_vigilancia, numero_inscripcion_clinica,
                   numero_factura, codigo_empleado, tipo_movimiento
            FROM registros 
            WHERE farmacia = ?
            ORDER BY fecha_ingreso DESC
        ''', ('farmacia1',))
        
        registros = cursor.fetchall()
        conn.close()
        
        print(f"📋 Se encontraron {len(registros)} registros para exportar")
        
        # Crear archivo Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Registros de Farmacia"
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Encabezados
        headers = ['Fecha Ingreso', 'Medicamento', 'Laboratorio', 'Cantidad', 'Fecha', 'Tipo Movimiento', 'N° Factura/Devolución', 'Fecha Vencimiento', 'Lote', 'Médico', 'Junta Vigilancia', 'Número Inscripción Clínica', 'Código Empleado', 'Observaciones']
        
        print("\n📋 Encabezados del Excel:")
        for i, header in enumerate(headers, 1):
            print(f"  Columna {i}: {header}")
            cell = ws.cell(row=1, column=i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Datos
        print(f"\n📊 Procesando {len(registros)} registros...")
        for row, registro in enumerate(registros, 2):
            print(f"\n  Registro {row-1}:")
            print(f"    - Laboratorio: {registro[0]}")
            print(f"    - Medicamento: {registro[1]}")
            print(f"    - Cantidad: {registro[2]}")
            print(f"    - Fecha: {registro[3]}")
            print(f"    - Tipo: {registro[13] if len(registro) > 13 else 'VENTA'}")
            print(f"    - N° Factura/Devolución: {registro[11] if len(registro) > 11 else ''}")
            
            # Formatear fecha de ingreso
            fecha_ingreso = registro[4]
            if fecha_ingreso:
                try:
                    fecha_obj = datetime.fromisoformat(fecha_ingreso.replace('Z', '+00:00'))
                    fecha_formateada = fecha_obj.strftime('%d/%m/%Y %H:%M:%S')
                except:
                    fecha_formateada = fecha_ingreso
            else:
                fecha_formateada = ''
            
            # Mapear datos a columnas
            ws.cell(row=row, column=1, value=fecha_formateada)  # Fecha Ingreso
            ws.cell(row=row, column=2, value=registro[1])  # Medicamento
            ws.cell(row=row, column=3, value=registro[0])  # Laboratorio
            ws.cell(row=row, column=4, value=registro[2])  # Cantidad
            ws.cell(row=row, column=5, value=registro[3])  # Fecha
            ws.cell(row=row, column=6, value=registro[13] if len(registro) > 13 else 'VENTA')  # Tipo Movimiento
            ws.cell(row=row, column=7, value=registro[11] if len(registro) > 11 else '')  # N° Factura/Devolución
            ws.cell(row=row, column=8, value=registro[6] if len(registro) > 6 else '')  # Fecha Vencimiento
            ws.cell(row=row, column=9, value=registro[7] if len(registro) > 7 else '')  # Lote
            ws.cell(row=row, column=10, value=registro[8] if len(registro) > 8 else '')  # Médico
            ws.cell(row=row, column=11, value=registro[9] if len(registro) > 9 else '')  # Junta Vigilancia
            ws.cell(row=row, column=12, value=registro[10] if len(registro) > 10 else '')  # Número Inscripción Clínica
            ws.cell(row=row, column=13, value=registro[12] if len(registro) > 12 else '')  # Código Empleado
            ws.cell(row=row, column=14, value=registro[5])  # Observaciones
        
        # Ajustar ancho de columnas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Guardar archivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"test_registros_farmacia_{timestamp}.xlsx"
        filepath = os.path.join(tempfile.gettempdir(), filename)
        wb.save(filepath)
        
        print(f"\n✅ Archivo Excel creado: {filepath}")
        print(f"📁 Ruta completa: {os.path.abspath(filepath)}")
        
        return filepath
        
    except Exception as e:
        print(f"❌ Error al exportar Excel: {e}")
        return None

def verify_excel_structure(filepath):
    """Verificar la estructura del archivo Excel generado"""
    print(f"\n🔍 Verificando estructura del archivo Excel: {filepath}")
    
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.active
        
        print(f"📊 Hoja: {ws.title}")
        print(f"📏 Filas: {ws.max_row}")
        print(f"📏 Columnas: {ws.max_column}")
        
        # Verificar encabezados
        print("\n📋 Encabezados encontrados:")
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            print(f"  Columna {col}: {header}")
        
        # Verificar datos
        print(f"\n📊 Datos de las primeras {min(3, ws.max_row-1)} filas:")
        for row in range(2, min(5, ws.max_row + 1)):
            print(f"\n  Fila {row}:")
            for col in range(1, ws.max_column + 1):
                value = ws.cell(row=row, column=col).value
                header = ws.cell(row=1, column=col).value
                print(f"    {header}: {value}")
        
        wb.close()
        return True
        
    except Exception as e:
        print(f"❌ Error al verificar Excel: {e}")
        return False

def main():
    """Función principal"""
    print("🧪 PRUEBA DE EXPORTACIÓN DE EXCEL")
    print("=" * 50)
    
    # Insertar datos de prueba
    if not insert_test_data():
        return
    
    # Probar exportación
    filepath = test_excel_export()
    if not filepath:
        return
    
    # Verificar estructura
    if not verify_excel_structure(filepath):
        return
    
    print("\n🎉 ¡Prueba completada exitosamente!")
    print("El archivo Excel se generó correctamente con la estructura esperada.")

if __name__ == "__main__":
    main() 