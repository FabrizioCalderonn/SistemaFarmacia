#!/usr/bin/env python3
"""
Script simplificado para procesar Excel sin pandas
"""
import os
import csv
from datetime import datetime
from openpyxl import load_workbook

# Importar funciones de la aplicación principal
from app import get_db_connection, DATABASE_TYPE

def backup_antes_de_procesar():
    """Crear backup antes de procesar Excel"""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener todos los registros
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('SELECT * FROM registros')
        else:
            cursor.execute('SELECT * FROM registros')
        
        registros = cursor.fetchall()
        conn.close()
        
        # Crear backup
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = f"backup_antes_excel_{timestamp}.csv"
        
        with open(backup_file, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(['id', 'medicamento', 'cantidad', 'precio', 'fecha', 'fecha_ingreso', 
                           'observaciones', 'laboratorio', 'tipo_movimiento', 'fecha_vencimiento', 
                           'lote', 'medico', 'junta_vigilancia', 'numero_inscripcion_clinica'])
            writer.writerows(registros)
        
        print(f"Backup creado: {backup_file}")
        return True
        
    except Exception as e:
        print(f"Error al crear backup: {e}")
        return False

def procesar_excel_original(ruta_archivo):
    """Procesar archivo Excel original"""
    try:
        if not os.path.exists(ruta_archivo):
            return False, "Archivo no encontrado"
        
        # Cargar archivo Excel
        wb = load_workbook(ruta_archivo)
        ws = wb.active
        
        # Conectar a la base de datos
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Limpiar inventario actual
        if DATABASE_TYPE == 'postgresql':
            cursor.execute('DELETE FROM inventario')
        else:
            cursor.execute('DELETE FROM inventario')
        
        # Procesar filas (asumiendo que la primera fila son encabezados)
        registros_insertados = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                if len(row) >= 5 and row[0] and row[1]:  # Verificar que hay datos
                    laboratorio = str(row[0]).strip()
                    medicamento = str(row[1]).strip()
                    presentacion = str(row[2]).strip() if len(row) > 2 else ''
                    precio = float(row[3]) if len(row) > 3 and row[3] else 0.0
                    stock = int(row[4]) if len(row) > 4 and row[4] else 0
                    
                    if DATABASE_TYPE == 'postgresql':
                        cursor.execute('''
                            INSERT INTO inventario (laboratorio, medicamento, presentacion, precio, stock)
                            VALUES (%s, %s, %s, %s, %s)
                        ''', (laboratorio, medicamento, presentacion, precio, stock))
                    else:
                        cursor.execute('''
                            INSERT INTO inventario (laboratorio, medicamento, presentacion, precio, stock)
                            VALUES (?, ?, ?, ?, ?)
                        ''', (laboratorio, medicamento, presentacion, precio, stock))
                    
                    registros_insertados += 1
                    
            except Exception as e:
                print(f"Error al procesar fila: {e}")
                continue
        
        conn.commit()
        conn.close()
        
        return True, f"Excel procesado exitosamente. {registros_insertados} registros insertados."
        
    except Exception as e:
        return False, f"Error al procesar Excel: {str(e)}"

if __name__ == "__main__":
    success, result = procesar_excel_original("inventarioraw.xlsx")
    print(result) 